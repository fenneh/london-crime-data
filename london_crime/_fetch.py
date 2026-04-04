"""Download, parse, and convert London Datastore resources to Parquet."""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Any

import httpx
import polars as pl

from ._sources import CKAN_BASE

logger = logging.getLogger(__name__)

DATA_DIR = Path(__file__).parent.parent / "data"


# ── CKAN API ───────────────────────────────────────────────────────────────────

def get_resources(short_id: str) -> list[dict[str, Any]]:
    """Fetch resources for a London Datastore dataset via its short ID."""
    url = f"{CKAN_BASE}/package_show?id={short_id}"
    r = httpx.get(url, timeout=30, follow_redirects=True)
    r.raise_for_status()
    body = r.json()
    if not body.get("success"):
        raise ValueError(f"CKAN returned success=false for {short_id!r}")
    return body["result"]["resources"]


def pick_resource(
    resources: list[dict[str, Any]],
    name_contains: str,
    prefer_historical: bool,
) -> dict[str, Any] | None:
    """Pick the best (most recent) matching resource."""
    historical_marker = "(Historical)"

    def matches(r: dict) -> bool:
        name = r.get("name", "")
        url  = r.get("url", "")
        # Must be a CSV or Excel
        if not any(url.lower().endswith(ext) for ext in (".csv", ".xlsx", ".xls")):
            if r.get("format", "").lower() not in ("csv", "xlsx", "xls"):
                return False
        # Must match name hint
        if name_contains and name_contains.lower() not in name.lower():
            return False
        # Historical filter
        if prefer_historical and historical_marker not in name:
            return False
        if not prefer_historical and historical_marker in name:
            return False
        return True

    candidates = [r for r in resources if matches(r)]
    if not candidates:
        return None

    # CKAN returns resources newest-first; take the first match.
    return candidates[0]


# ── Download + parse ───────────────────────────────────────────────────────────

def download_url(url: str) -> bytes:
    logger.info("  Downloading: %s", url)
    r = httpx.get(url, timeout=300, follow_redirects=True)
    r.raise_for_status()
    return r.content


def parse_csv(data: bytes) -> pl.DataFrame:
    for enc in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return pl.read_csv(
                io.BytesIO(data),
                infer_schema_length=10_000,
                ignore_errors=True,
                truncate_ragged_lines=True,
                encoding=enc,
            )
        except Exception:
            continue
    raise ValueError("Could not parse CSV in any encoding")


def parse_excel(data: bytes) -> pl.DataFrame:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    frames = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        data_rows = [list(r) for r in rows[1:] if any(v is not None for v in r)]
        if not data_rows:
            continue
        df = pl.DataFrame(
            {h: [r[i] if i < len(r) else None for r in data_rows] for i, h in enumerate(headers)}
        )
        df = df.with_columns(pl.lit(sheet).alias("_sheet"))
        frames.append(df)

    if not frames:
        raise ValueError("No data sheets found in Excel file")
    if len(frames) == 1:
        return frames[0]
    return _concat(frames)


def parse(data: bytes, url: str) -> pl.DataFrame:
    url_lower = url.lower()
    if url_lower.endswith((".xlsx", ".xls")):
        return parse_excel(data)
    return parse_csv(data)


# ── Normalise ──────────────────────────────────────────────────────────────────

def normalise(df: pl.DataFrame) -> pl.DataFrame:
    renamed = {}
    for col in df.columns:
        new = (
            col.strip()
            .lower()
            .replace(" ", "_")
            .replace("-", "_")
            .replace("/", "_")
            .replace("(", "")
            .replace(")", "")
            .replace(".", "_")
        )
        while "__" in new:
            new = new.replace("__", "_")
        renamed[col] = new.strip("_")
    df = df.rename(renamed)

    # Cast obvious numeric string columns
    casts = []
    for col, dtype in zip(df.columns, df.dtypes):
        if dtype == pl.Utf8:
            sample = df[col].drop_nulls().head(100)
            if len(sample) == 0:
                continue
            non_empty = sample.filter(sample != "")
            if len(non_empty) == 0:
                continue
            try:
                non_empty.cast(pl.Int64)
                casts.append(pl.col(col).cast(pl.Int64, strict=False))
                continue
            except Exception:
                pass
            try:
                non_empty.cast(pl.Float64)
                casts.append(pl.col(col).cast(pl.Float64, strict=False))
            except Exception:
                pass
    if casts:
        df = df.with_columns(casts)
    return df


def _concat(frames: list[pl.DataFrame]) -> pl.DataFrame:
    # Build unified column→dtype map; where dtypes conflict, fall back to String
    all_cols: dict[str, pl.DataType] = {}
    for df in frames:
        for col, dtype in zip(df.columns, df.dtypes):
            if col not in all_cols:
                all_cols[col] = dtype
            elif all_cols[col] != dtype:
                all_cols[col] = pl.String

    aligned = []
    for df in frames:
        casts = []
        for col, target in all_cols.items():
            if col in df.columns and df[col].dtype != target:
                casts.append(pl.col(col).cast(pl.String).alias(col))
        if casts:
            df = df.with_columns(casts)
        missing = {c: t for c, t in all_cols.items() if c not in df.columns}
        if missing:
            df = df.with_columns([pl.lit(None).cast(t).alias(c) for c, t in missing.items()])
        aligned.append(df.select(list(all_cols.keys())))
    return pl.concat(aligned, rechunk=True)


# ── Save ───────────────────────────────────────────────────────────────────────

def save(df: pl.DataFrame, output: str) -> Path:
    path = DATA_DIR / output
    path.parent.mkdir(parents=True, exist_ok=True)
    df.write_parquet(path, compression="zstd")
    size_kb = path.stat().st_size // 1024
    logger.info("  Saved %s (%d KB, %d rows × %d cols)", output, size_kb, len(df), len(df.columns))
    return path


# ── High-level refresh ─────────────────────────────────────────────────────────

def _filter_resources(
    resources: list[dict[str, Any]],
    name_contains: str,
    prefer_historical: bool,
) -> list[dict[str, Any]]:
    """Return all resources matching the source filters (no sorting)."""
    historical_marker = "(Historical)"

    def matches(r: dict) -> bool:
        name = r.get("name", "")
        url = r.get("url", "")
        if not any(url.lower().endswith(ext) for ext in (".csv", ".xlsx", ".xls")):
            if r.get("format", "").lower() not in ("csv", "xlsx", "xls"):
                return False
        if name_contains and name_contains.lower() not in name.lower():
            return False
        if prefer_historical and historical_marker not in name:
            return False
        if not prefer_historical and historical_marker in name:
            return False
        return True

    return [r for r in resources if matches(r)]


def refresh_source(source: dict) -> dict[str, int]:
    """Refresh one source. Returns {output_path: row_count}."""
    logger.info("Refreshing: %s", source["name"])

    resources = get_resources(source["short_id"])
    logger.info("  %d resources available in dataset", len(resources))

    if source.get("combine_resources"):
        return _refresh_combined(source, resources)

    resource = pick_resource(
        resources,
        name_contains=source.get("resource_name_contains", ""),
        prefer_historical=source.get("prefer_historical", False),
    )
    if resource is None:
        logger.warning("  No matching resource found — skipping")
        logger.warning(
            "  Available resources: %s",
            [r.get("name") for r in resources[:10]],
        )
        return {}

    logger.info("  Selected: %s", resource.get("name", "?"))
    logger.info("  URL: %s", resource.get("url", "?"))

    data = download_url(resource["url"])
    df = parse(data, resource["url"])
    df = normalise(df)

    path = save(df, source["output"])
    return {str(path): len(df)}


def _refresh_combined(source: dict, resources: list[dict[str, Any]]) -> dict[str, int]:
    """Download multiple rolling-window snapshots and combine into one file.

    Uses stride sampling to avoid downloading every resource for large datasets.
    stride=1 downloads everything; stride=N downloads every N-th resource.
    Always includes the first (newest) and last (oldest) resource.
    """
    candidates = _filter_resources(
        resources,
        name_contains=source.get("resource_name_contains", ""),
        prefer_historical=source.get("prefer_historical", False),
    )
    if not candidates:
        logger.warning("  No matching resources found — skipping")
        return {}

    stride = source.get("combine_stride", 1)
    indices = list(range(0, len(candidates), stride))
    if (len(candidates) - 1) not in indices:
        indices.append(len(candidates) - 1)

    logger.info(
        "  Combining %d/%d resources (stride=%d)",
        len(indices), len(candidates), stride,
    )

    frames = []
    for i in indices:
        res = candidates[i]
        url = res.get("url", "")
        try:
            data = download_url(url)
            df = parse(data, url)
            df = normalise(df)
            frames.append(df)
            logger.info("  Resource #%d: %d rows", i, len(df))
        except Exception as exc:
            logger.warning("  Skipping resource #%d (%s): %s", i, url, exc)

    if not frames:
        logger.warning("  All resources failed — skipping")
        return {}

    combined = _concat(frames).unique()
    logger.info("  Combined: %d rows after deduplication", len(combined))

    path = save(combined, source["output"])
    return {str(path): len(combined)}
